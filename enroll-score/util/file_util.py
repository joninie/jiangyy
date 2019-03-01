#!/usr/bin/env python
# coding=utf-8

import xlrd, xlwt
from openpyxl import Workbook


def readXLS_R1(filename):
    ''' 使用xlrd读取xls，如果文件是使用xlwt写入的请用这个函数读取，只要第一个sheet '''
    book = xlrd.open_workbook(filename)
    sheets = book.sheets()
    content = [[[str(int(c.value)) if type(c.value) == float and c.value == int(c.value) else str(c.value) for c in sht.row(i)] for i in range(sht.nrows)] for sht in sheets]
    return content[0]


def writeXLS(content, filename):
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheet1')

    for i, cont in enumerate(content):
        for j, c in enumerate(cont):
            ws.write(i, j, c)
    wb.save(filename)


def writeXLSX(content, filename):
    wb = Workbook()
    ws = wb.active
    for i, cont in enumerate(content):
        for j, c in enumerate(cont):
            ws.cell(i + 1, j + 1, c)

    wb.save(filename)